#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Сборка XLSX-базы контактов из research.json.

Использование:
    python make_xlsx.py research.json out.xlsx

Схема research.json описана в references/output-formats.md.
Требует openpyxl:  pip install openpyxl
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COMPANY_COLS = [
    ("step", "Ступень"),
    ("tier", "Тир"),
    ("name", "Компания"),
    ("best_contact", "Лучший контакт"),
    ("step_note", "Чем хорош / чего не хватает"),
    ("trigger", "Повод для контакта"),
    ("inn", "ИНН"),
    ("ogrn", "ОГРН"),
    ("legal_address", "Юр. адрес"),
    ("actual_address", "Факт. адрес"),
    ("site", "Сайт"),
    ("okved", "Отрасль / ОКВЭД"),
    ("revenue", "Выручка"),
    ("headcount", "Численность"),
    ("director", "Директор"),
    ("founders", "Учредители"),
    ("related_entities", "Связанные юрлица"),
    ("general_phone", "Общий телефон"),
    ("general_email", "Общая почта"),
    ("socials", "Соцсети"),
    ("status", "Статус проверки"),
    ("sources", "Источники"),
]

CONTACT_COLS = [
    ("company", "Компания"),
    ("priority", "Приоритет"),
    ("fio", "ФИО"),
    ("position", "Должность"),
    ("role", "Роль"),
    ("phone", "Телефон"),
    ("email", "Email"),
    ("telegram", "Telegram"),
    ("social", "Соцсеть"),
    ("confidence", "Confidence"),
    ("source_date", "Дата источника"),
    ("source", "Источник"),
    ("comment", "Комментарий"),
]

CONF_FILL = {
    "подтверждён": PatternFill("solid", fgColor="C6EFCE"),
    "вероятен": PatternFill("solid", fgColor="FFEB9C"),
    "гипотеза": PatternFill("solid", fgColor="FFC7CE"),
}

TIER_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="FFEB9C"),
    "C": PatternFill("solid", fgColor="E7E6E6"),
}

NUMERIC = {"step", "priority", "rank"}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def flat(value):
    """Любую вложенную структуру превращает в читаемую строку ячейки."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        return "; ".join(f"{k}: {flat(v)}" for k, v in value.items() if v)
    if isinstance(value, list):
        return "\n".join(flat(v) for v in value if v)
    return str(value)


def write_sheet(ws, cols, rows):
    ws.append([title for _, title in cols])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([
            row.get(key) if key in NUMERIC and isinstance(row.get(key), (int, float))
            else flat(row.get(key))
            for key, _ in cols
        ])

    def paint(col_key, palette):
        idx = next((i for i, (k, _) in enumerate(cols, start=1) if k == col_key), None)
        if not idx:
            return
        for r in range(2, ws.max_row + 1):
            fill = palette.get(ws.cell(r, idx).value)
            if fill:
                ws.cell(r, idx).fill = fill

    paint("confidence", CONF_FILL)
    paint("tier", TIER_FILL)

    for i, _ in enumerate(cols, start=1):
        longest = max(
            (len(str(ws.cell(r, i).value or "").split("\n")[0]) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 12), 55)

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(cols) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)

    def as_num(value, default=0):
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    # Порядок: явный rank, если модель его проставила; иначе ступень 1 -> 6.
    # Записи без ступени уходят в самый низ, а не наверх.
    companies = sorted(
        data.get("companies", []),
        key=lambda c: (
            as_num(c.get("rank"), 10_000),
            as_num(c.get("step"), 99),
            c.get("name", ""),
        ),
    )

    contacts = []
    for comp in companies:
        # Внутри компании — по приоритету (1 = лучший вход на ЛПР); без приоритета в хвост.
        for contact in sorted(
            comp.get("contacts", []),
            key=lambda k: as_num(k.get("priority"), 999),
        ):
            row = dict(contact)
            row["company"] = comp.get("name", "")
            contacts.append(row)

    wb = Workbook()
    write_sheet(wb.active, COMPANY_COLS, companies)
    wb.active.title = "Компании"
    write_sheet(wb.create_sheet("Контакты"), CONTACT_COLS, contacts)

    not_found = data.get("not_found", [])
    if not_found:
        write_sheet(
            wb.create_sheet("Не найдено"),
            [("company", "Компания"), ("reason", "Причина")],
            not_found,
        )

    wb.save(sys.argv[2])
    print(f"OK: {sys.argv[2]} — компаний {len(companies)}, контактов {len(contacts)}, не найдено {len(not_found)}")


if __name__ == "__main__":
    main()
